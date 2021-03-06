from app import db
from flask_login import current_user, login_required
from app.main import bp
from app.models import User, UserRoles, Ecwid, OrderApproval, EventLog, EventType, OrderStatus, Location
from flask import render_template, redirect, url_for, flash, jsonify, current_app, Response
from app.main.forms import OrderCommentsForm, OrderApprovalForm, ChangeQuantityForm, ChangeLocationForm
from datetime import datetime, timezone
from app.ecwid import EcwidAPIException
from app.main.utils import DATE_TIME_FORMAT, role_required, ecwid_required, PrepareOrder, GetProductApproval, role_required_ajax, ecwid_required_ajax, role_forbidden, role_forbidden_ajax, SendEmailNotification

from openpyxl import load_workbook
from copy import copy
from openpyxl.writer.excel import save_virtual_workbook

'''
################################################################################
Consts
################################################################################
'''
_DISALLOWED_ORDERS_ITEM_FIELDS = ['productId', 'id', 'categoryId']
_DISALLOWED_ORDERS_FIELDS = ['vendorOrderNumber', 'customerId', 'externalFulfillment', 'createDate', 'initiative', 'updateDate', 'reviewers', 'events', 'status', 'positions']

'''
################################################################################
Approve page
################################################################################
'''

def GetOrder(order_id):
	try:
		response = current_user.hub.GetStoreOrders(orderNumber = order_id)
		if 'items' not in response or len(response['items']) == 0:
			raise EcwidAPIException('Такой заявки не существует')
		order = response['items'][0]
		if PrepareOrder(order) is False:
			raise EcwidAPIException('Заявка не принадлежит ни одному из инициаторов')
		if current_user.role == UserRoles.initiative and order['email'] != current_user.email:
			raise EcwidAPIException('Вы не являетесь владельцем этой заявки')
	except EcwidAPIException as e:
		flash('Ошибка API: {}'.format(e))
		order = None
	return order
	
@bp.route('/order/<int:order_id>')
@login_required
@role_forbidden([UserRoles.default])
@ecwid_required
def ShowOrder(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	
	vendors = Ecwid.query.filter(Ecwid.ecwid_id == current_user.ecwid_id).all()
	vendors = {str(vendor.store_id):vendor.store_name for vendor in vendors}
	
	for product in order['items']:
		try:
			dash = product['sku'].index('-')
			product['vendor'] = vendors[product['sku'][:dash]]
		except (ValueError, KeyError):
			product['vendor'] = ''
		
	approval_form = OrderApprovalForm()
	quantity_form = ChangeQuantityForm()
	comment_form = OrderCommentsForm()
	
	location = Location.query.filter(Location.name.ilike(order['refererId'])).first()
	locations = Location.query.order_by(Location.name).all()
	location_form = ChangeLocationForm()
	location_form.location_name.choices = [(l.id, l.name) for l in locations]
	if location is None:
		location_form.location_name.choices.append((0, 'Выберите площадку...'))
		location_form.location_name.default = 0
	else:
		location_form.location_name.default = location.id
	location_form.process()
	
	return render_template('approve.html',
							order = order,
							location_form = location_form,
							comment_form = comment_form,
							approval_form = approval_form,
							quantity_form = quantity_form)

@bp.route('/comment/<int:order_id>', methods=['POST'])
@login_required
@role_forbidden([UserRoles.admin, UserRoles.default])
def SaveComment(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	form = OrderCommentsForm()
	if form.validate_on_submit():
		stripped = form.comment.data.strip()
		if len(stripped) > 0:
			comment = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.comment, data=stripped, timestamp = datetime.now(tz = timezone.utc))
			db.session.add(comment)
			flash('Комментарий успешно добавлен')
		else:
			flash('Комментарий не может быть пустым')
		db.session.commit()
	return redirect(url_for('main.ShowOrder', order_id = order_id))
	
@bp.route('/location/<int:order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.initiative])
def SaveLocation(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	locations = Location.query.order_by(Location.name).all()
	form = ChangeLocationForm()
	form.location_name.choices = [(l.id, l.name) for l in locations]
	if form.validate_on_submit():
		referer_id = dict(form.location_name.choices).get(form.location_name.data)
		try:
			response = current_user.hub.UpdateStoreOrder(order_id, {'refererId':referer_id})
			message = 'площадка была "{}", стала "{}"'.format(order['refererId'], referer_id)
			event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.modified, data=message, timestamp = datetime.now(tz = timezone.utc))
			db.session.add(event)
			db.session.commit()
			flash('Площадка успешно изменена.')
		except EcwidAPIException:
			flash('Не удалось присвоить данную площадку.')
	else:
		flash('Невозможно присвоить данную площадку.')
	return redirect(url_for('main.ShowOrder', order_id = order_id))


@bp.route('/approval/<int:order_id>', methods=['POST'])
@login_required
@role_forbidden([UserRoles.admin, UserRoles.default, UserRoles.initiative])
def SaveApproval(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	form = OrderApprovalForm()
	if form.validate_on_submit():
		try:
			order_approval = OrderApproval.query.filter(OrderApproval.order_id == order_id, OrderApproval.user_id == current_user.id, OrderApproval.product_id == None).first()
			if form.product_id.data is None:
				OrderApproval.query.filter(OrderApproval.order_id == order_id, OrderApproval.user_id == current_user.id).delete()
				if order_approval is None:
					order_approval = OrderApproval(order_id = order_id, product_id = None, user_id = current_user.id)
					db.session.add(order_approval)
					event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.approved, data='заявка', timestamp = datetime.now(tz = timezone.utc))
					if PrepareOrder(order):
						if order['status'] == OrderStatus.approved:
							SendEmailNotification('approved', order)
				else:
					event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.disapproved, data='согласованная ранее заявка', timestamp = datetime.now(tz = timezone.utc))
					for product in order['items']:
						product_approval = OrderApproval(order_id = order_id, product_id=product['id'], user_id = current_user.id)
						db.session.add(product_approval)
					if order['status'] != OrderStatus.not_approved:
						SendEmailNotification('disapproved', order)
			else:
				product_approval = OrderApproval.query.filter(OrderApproval.order_id == order_id, OrderApproval.user_id == current_user.id, OrderApproval.product_id == form.product_id.data).first()
				if product_approval is not None:
					db.session.delete(product_approval)
					message = 'товар <span class="product-sku text-primary">{}</span>'.format(product_approval.product_sku)
					event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.approved, data=message, timestamp = datetime.now(tz = timezone.utc))
				else:
					if order_approval is not None:
						db.session.delete(order_approval)
					product_approval = OrderApproval(order_id = order_id, product_id = form.product_id.data, user_id = current_user.id, product_sku = form.product_sku.data.strip())
					db.session.add(product_approval)
					message = 'товар <span class="product-sku text-primary">{}</span>'.format(product_approval.product_sku)
					event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.disapproved, data=message, timestamp = datetime.now(tz = timezone.utc))
					if order['status'] != OrderStatus.not_approved:
						SendEmailNotification('disapproved', order)
			db.session.add(event)
			db.session.commit()
		except EcwidAPIException as e:
			flash('Ошибка API: {}'.format(e))
	return redirect(url_for('main.ShowOrder', order_id = order_id))


@bp.route('/delete/<int:order_id>')
@login_required
@role_required([UserRoles.initiative])
@ecwid_required
def DeleteOrder(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	try:
		response = current_user.hub.DeleteStoreOrder(order_id = order_id)
		events = EventLog.query.join(User).filter(EventLog.order_id == order_id, User.ecwid_id == current_user.ecwid_id).all()
		approvals = OrderApproval.query.join(User).filter(OrderApproval.order_id == order_id, User.ecwid_id == current_user.ecwid_id).all()
		for approval in approvals:
			db.session.delete(approval)
		for event in events:
			db.session.delete(event)
		db.session.commit()
		flash('Заявка успешно удалена')
	except EcwidAPIException as e:
		flash('Ошибка API: {}'.format(e))
	return redirect(url_for('main.ShowIndex'))
	
@bp.route('/duplicate/<int:order_id>')
@login_required
@role_required([UserRoles.initiative])
@ecwid_required
def DuplicateOrder(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	for key in _DISALLOWED_ORDERS_FIELDS:
		order.pop(key, None)
	try:
		response = current_user.hub.SetStoreOrder(order)
		if 'id' not in response:
			raise EcwidAPIException('Не удалось дулировать заявку')
		message = 'Заявка дублирована с внутренним номером <a href={}>{}</a>'.format(url_for('main.ShowOrder', order_id = response['id']), response['id'])
		event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.duplicated, data=message, timestamp = datetime.now(tz = timezone.utc))
		db.session.add(event)
		message = 'Заявка дублирована из заявки <a href={}>{}</a>'.format(url_for('main.ShowOrder', order_id = order_id), order_id)
		event = EventLog(user_id = current_user.id, order_id = response['id'], type=EventType.duplicated, data=message, timestamp = datetime.now(tz = timezone.utc))
		db.session.add(event)
		db.session.commit()
		flash('Заявка успешно дублирована')
	except EcwidAPIException as e:
		flash('Ошибка API: {}'.format(e))
	return redirect(url_for('main.ShowOrder', order_id = order_id))


@bp.route('/quantity/<int:order_id>', methods=['POST'])
@login_required
@role_required([UserRoles.initiative])
@ecwid_required
def SaveQuantity(order_id):
	new_total = ''
	form = ChangeQuantityForm()
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	if form.validate_on_submit():
		for i, product in enumerate(order['items']):
			if form.product_id.data == product['id']:
				order['total'] += (form.product_quantity.data - product['quantity'])*product['price']
				if order['total'] < 0:
					order['total'] = 0
				new_total = '{:,.2f}'.format(order['total'])
				message = '<span class="product-sku text-primary">{}</span> количество было {} стало {}'.format(product['sku'], product['quantity'], form.product_quantity.data)
				product['quantity'] = form.product_quantity.data
				index = i
				break
		else:
			flash('Указанный товар не найден в заявке')
			return redirect(url_for('main.ShowOrder', order_id = order_id))

		approvals = OrderApproval.query.join(User).filter(OrderApproval.order_id == order_id, User.ecwid_id == current_user.ecwid_id).all()
		for approval in approvals:
			db.session.delete(approval)
		if form.product_quantity.data == 0:
			order['items'].pop(index)
		try:
			if len(order['items']) == 0:
				response = current_user.hub.DeleteStoreOrder(order_id = order_id)
				events = EventLog.query.join(User).filter(EventLog.order_id == order_id, User.ecwid_id == current_user.ecwid_id).all()
				for event in events:
					db.session.delete(event)
				flash('Заявка удалена, вернитесь на главную страницу')
				db.session.commit()
				return redirect(url_for('main.ShowIndex'))
			else:
			
				for key in _DISALLOWED_ORDERS_FIELDS:
					order.pop(key, None)
						
				response = current_user.hub.UpdateStoreOrder(order_id, order)
				event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.modified, data=message, timestamp=datetime.now(tz = timezone.utc))
				db.session.add(event)
				flash('Количество {} было изменено'.format(product['sku']))
			db.session.commit()
		except EcwidAPIException as e:
			flash('Ошибка API: {}'.format(e))
			db.session.rollback()
	else:
		for error in form.product_id.errors + form.product_quantity.errors:
			flash(error)
	return redirect(url_for('main.ShowOrder', order_id = order_id))

	
@bp.route('/process/<int:order_id>')
@login_required
@role_required([UserRoles.approver])
@ecwid_required
def ProcessHubOrder(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
		
	for key in _DISALLOWED_ORDERS_FIELDS:
		order.pop(key, None)
	for product in order['items']:
		for key in _DISALLOWED_ORDERS_ITEM_FIELDS:
			product.pop(key, None)

	stores = Ecwid.query.filter(Ecwid.ecwid_id == current_user.ecwid_id).all()
	got_orders = {}
	for store in stores:
		products = list()
		total = 0
		for product in order['items']:
			try:
				dash = product['sku'].index('-')
			except ValueError:
				continue
			if product['sku'][:dash] == str(store.store_id):
				product_new = product.copy()
				product_new['sku'] = product_new['sku'][dash+1:]
				products.append(product_new)
				total += product_new['price'] * product_new['quantity']
		if len(products) == 0:
			continue
		items = order['items']
		order['items'] = products
		order['subtotal'] = total
		order['total'] = total
		order['email'] = current_user.email
		try:
			result = store.SetStoreOrder(order)
			got_orders[store.store_name] = result['id']
		except EcwidAPIException as e:
			flash('Не удалось перезаказать товары у {}'.format(store.store_name))
		order['items'] = items

	if len(got_orders) > 0:
		message = ', '.join(f'{vendor} (#{order})' for vendor,order in got_orders.items())
		try:
			referer = current_user.name if current_user.name else current_user.email
			current_user.hub.UpdateStoreOrder(order_id, {'externalFulfillment':True})
			event = EventLog(user_id = current_user.id, order_id = order_id, type=EventType.vendor, data=message, timestamp=datetime.now(tz = timezone.utc))
			db.session.add(event)
			db.session.commit()
		except EcwidAPIException as e:
			flash('Ошибка API: {}'.format(e))
			flash('Не удалось сохранить информацию о передаче поставщикам')
		flash('Заявка была отправлена поставщикам: {}'.format(message))
	else:
		flash('Не удалось перезаказать данные товары у зарегистрованных поставщиков')

	return redirect(url_for('main.ShowOrder', order_id = order_id))
	
	
@bp.route('/notify/<int:order_id>')
@login_required
@role_required_ajax([UserRoles.initiative])
@ecwid_required_ajax
def NotifyApprovers(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	SendEmailNotification('modified', order)
	flash('Уведомление успешно выслано')
	return redirect(url_for('main.ShowOrder', order_id = order_id))
	

@bp.route('/report/<int:order_id>')
@login_required
@role_forbidden([UserRoles.default])
@ecwid_required
def GetExcelReport(order_id):
	order = GetOrder(order_id)
	if order is None:
		return redirect(url_for('main.ShowIndex'))
	
	vendors = Ecwid.query.filter(Ecwid.ecwid_id == current_user.ecwid_id).all()
	vendors = {str(vendor.store_id):vendor.store_name for vendor in vendors}
	
	data_len = len(order['items'])
	starting_row = 11
	wb = load_workbook(filename = 'template.xlsx')
	ws = wb.active
	ws['P17'] = order['initiative'].name
	if data_len > 1:
		for merged_cell in ws.merged_cells.ranges:
			if merged_cell.bounds[1] >= starting_row:
				merged_cell.shift(0, data_len)
		ws.insert_rows(starting_row, data_len-1)
	for k,i in enumerate(range(starting_row, starting_row+data_len)):
		product = order['items'][k]
		ws.row_dimensions[i].height = 50
		if data_len > 1:
			for j in range(1, 20):
				target_cell = ws.cell(row=i, column=j)
				source_cell = ws.cell(row=starting_row + data_len - 1, column=j)
				target_cell._style = copy(source_cell._style)
				target_cell.font = copy(source_cell.font)
				target_cell.border = copy(source_cell.border)
				target_cell.fill = copy(source_cell.fill)
				target_cell.number_format = copy(source_cell.number_format)
				target_cell.protection = copy(source_cell.protection)
				target_cell.alignment = copy(source_cell.alignment)
		ws.cell(i, 1).value = k + 1
		ws.cell(i, 5).value = product['name']
		try:
			dash = product['sku'].index('-')
			vendor = vendors[product['sku'][:dash]]
		except (ValueError, KeyError):
			vendor = ''
		ws.cell(i, 3).value = order['refererId']
		ws.cell(i, 7).value = vendor
		ws.cell(i, 8).value = product['quantity']
		c1 = ws.cell(i, 8).coordinate
		ws.cell(i, 10).value = product['price']
		c2 = ws.cell(i, 10).coordinate
		ws.cell(i, 12).value = f"={c1}*{c2}"
		
	data = save_virtual_workbook(wb)
	return Response (data, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers = {'Content-Disposition':'attachment;filename=report.xlsx'})
	
	

